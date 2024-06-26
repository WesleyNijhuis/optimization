import os

import numpy as np

wd = '/Users/Wesley/Downloads'


import gurobipy as gp
options = {
    "WLSACCESSID":"19417b59-15af-4762-93dd-83b48b1ec4a9",
    "WLSSECRET":"f288dc19-ff29-4ff7-9cab-d05954abe26b",
    "LICENSEID":2507232
}
with gp.Env(params=options) as env, gp.Model(env=env) as model:
    # Formulate problem
    model.optimize()


np.set_printoptions(threshold=np.inf)

from gurobipy import Model, GRB, quicksum
from openpyxl import load_workbook
from tabulate import tabulate
import numpy as np
import csv
import pandas as pd
import xlsxwriter
import matplotlib.pyplot as plt





BM = Model('Basic_Model')


wb = load_workbook(os.path.join(wd,'Netwerk 2.xlsx'))

ws = wb.active


B = np.array([i[0].value for i in wb.worksheets[0]['A3':'A20']])
V = np.array([i[0].value for i in wb.worksheets[2]['A2':'A5']])
T = np.array([i[0].value for i in wb.worksheets[1]['A2':'A6']])
N = np.array([i[0].value for i in wb.worksheets[1]['A7':'A11']])
R = np.append(T,N) #NEW geen T maar R(={T,N})
P = np.array([i[0].value for i in wb.worksheets[3]['A3':'A26']])

## Batteries
cb = np.array([i[0].value for i in wb.worksheets[0]['B3':'B20']])  # Investment cost [Euro]
cap = np.array([i[0].value for i in wb.worksheets[0]['C3':'C20']])  # Capacity [kW]
mbl = 0.1  # Min battery level at all times
Lb0 = np.array([i[0].value for i in wb.worksheets[0]['D3':'D20']])  # Initial values: battery level [kW]

ut_aux = np.array([i[0].value for i in wb.worksheets[0]['E3':'E20']])  # Initial values: battery location at terminal
Ut0 = np.zeros((len(B), len(T)))
for b in range(len(B)):
    if ut_aux[b] > 0:
        Ut0[b, ut_aux[b] - 1] = 1

uv_aux = np.array([i[0].value for i in wb.worksheets[0]['F3':'F20']])  # Initial values: battery location on vessel
Uv0 = np.zeros((len(B), len(V)))
for b in range(len(B)):
    if uv_aux[b] > 0:
        Uv0[b, uv_aux[b] - 1] = 1

k_aux = np.array([i[0].value for i in wb.worksheets[0]['G3':'G20']])  # Initial values: battery usage on terminal
k0 = np.zeros((len(B), len(T)))
for b in range(len(B)):
    if k_aux[b] > 0:
        k0[b, k_aux[b] - 1] = 1

m_aux = np.array([i[0].value for i in wb.worksheets[0]['H3':'H20']])  # Initial values: battery usage on vessel

m0 = np. zeros (( len (B),len (V)))
for b in range ( len (B)):
    if m_aux [b] >0:
        m0[b, m_aux [b ] -1]=1

## Vessels
chv = np. array ([i [0]. value for i in wb. worksheets [2]['B2':'B5']]) # Total number of charging points in vessel

sp_aux = [np. array ([i [0]. value for i in wb. worksheets [3]['B3':'B26']]) , # Sailing profile of vessels
    np. array ([i [0]. value for i in wb. worksheets [3]['C3':'C26']]) ,
    np. array ([i [0]. value for i in wb. worksheets [3]['D3':'D26']]) ,
    np. array ([i [0]. value for i in wb. worksheets [3]['E3':'E26']]) ]

sp_aux = np.array(sp_aux)

sp = np.zeros((len(V), len(R), len(P)))
for v in range(len(V)):
    for p in range(len(P)):
        if sp_aux[v, p] > 0:
            sp[v, sp_aux[v, p] - 1, p] = 1

sp_aux_var = np.zeros((len(V), len(P)))
for v in range(len(V)):
    sp_aux_var[v,0] = sp_aux[v,0]


pv_aux = [np.array([i[0].value for i in wb.worksheets[3]['C3':'C26']], dtype=float),  # Power requirement per time step
           np.array([i[0].value for i in wb.worksheets[3]['E3':'E26']], dtype=float),
           np.array([i[0].value for i in wb.worksheets[3]['G3':'G26']], dtype=float),
           np.array([i[0].value for i in wb.worksheets[3]['I3':'I26']], dtype=float)]

pv = np.array(pv_aux)

#NEW reading max pv for later use
pv_max = np.array([i[0].value for i in wb.worksheets[2]['E2':'E5']])
print(pv_max)

#NEW reading top speed for later use
ts = np.array([i[0].value for i in wb.worksheets[2]['D2':'D5']])

print(ts)

mlv = 1  # Min battery level to be placed on vessel

## Docking Stations
cds = np.array([i[0].value for i in wb.worksheets[1]['B2':'B6']])  # Investment cost [Euro]
cht = np.array([i[0].value for i in wb.worksheets[1]['C2':'C6']])  # Total number of charging points at terminal
pw = np.array([i[0].value for i in wb.worksheets[1]['D2':'D6']])  # Power grids provide the DS [kWh]

if len(cds) != len(T):
    print("Error: Length of 'cds' does not match the length of 'T'")

## NEW grid specifications

CC_p = np.array([i[0].value for i in wb.worksheets[4]['B2:B25']])

if len(CC_p) != len(P):
    print("Error: Length of 'CC_p' does not match the length of 'P'")

pw_max = max(pw)

#Investment horizon
jaren = 10
beta = 365.25 * jaren / ((len(P)/24)) #horizon over X aantal jaar (in dagen, zeilprofiel van een week)


## NEW - Mapping
LT = np.zeros((len(V), len(R), len(P))) # Logistics Tensor - destinations and maximum arrival times
for v in range(len(V)):
    for p in range(len(P)):
        if sp_aux[v, p] > 0:
            if sp_aux[v,p] != sp_aux[v, p - 1]:
                LT[v, sp_aux[v, p] - 1, p] = 1

DM = [] #Distance Matrix
for row in range(2,12):
    current_row = []
    for col in range(2,12):
        current_cell = wb.worksheets[6].cell(row=row, column=col).value
        current_row.append(current_cell)
    DM.append(current_row)
DM = np.array(DM)
DM_max = np.max(DM)

## Others
M1 = 2 * len(P)  # Large number
M2 = 1 / max(pv_max)  # Small number (NEW definition)
M3 = DM_max #NEW large number for distance = max(ts[v]) * len(P)
M4 = max(cap) #large number for battery charge
M5 = 1 / pw_max #small number for battery charging

#VARIABLES
# x[b,t,p]: Binary variable if battery is at ds
x = {}
for b in range(len(B)):
    for t in range(len(T)):
        for p in range(len(P)):
            x[b, t, p] = BM.addVar(lb=0, vtype=GRB.BINARY, name="x[" + str(b) + "," + str(t) + "," + str(p) + "]")

# y[b,v,p]: Binary variable if battery is at a vessel
y = {}
for b in range(len(B)):
    for v in range(len(V)):
        for p in range(len(P)):
            y[b,v,p] = BM.addVar(lb=0, vtype=GRB.BINARY, name="y[" + str(b) + "," + str(v) + "," + str(p) + "]")

# n[b]: Binary variable if a battery is used
n = {}
for b in range(len(B)):
    n[b] = BM.addVar(lb=0, vtype=GRB.BINARY, name="n[" + str(b) + "]")

# u[t]: Binary variable if a terminal is used as DS
u = {}
for t in range(len(T)):
    u[t] = BM.addVar(lb=0, vtype=GRB.BINARY, name="u[" + str(t) + "]")

# l[b,p]: Continuous variable for battery level
l = {}
for b in range(len(B)):
    for p in range(len(P)):
        l[b, p] = BM.addVar(lb=0, ub = cap[b], vtype=GRB.CONTINUOUS, name="l[" + str(b) + "," + str(p) + "]")

# m[b,v,p]: Binary variable if battery being used on a vessel
m = {}
for b in range(len(B)):
    for v in range(len(V)):
        for p in range(len(P)):
            m[b, v, p] = BM.addVar(lb=0, vtype=GRB.BINARY, name="m[" + str(b) + "," + str(v) + "," + str(p) + "]")


# k[b,t,p]: Binary variable if battery being charged at a terminal
k = {}
for b in range ( len (B)):
    for t in range ( len (T)):
        for p in range ( len (P)):
            k[b,t,p]= BM. addVar (lb =0, vtype = GRB . BINARY , name ="k["+str(b)+","+ str(t)+","+ str(p)+"]")

# NEW variable charging to better allow for optimization
pw_var = {}

for b in range(len (B)):
    for t in range(len(T)):
        for p in range(len(P)):
            pw_var[b,t,p] = BM.addVar(lb=0, ub = pw[t], vtype=GRB.CONTINUOUS, name="pw_var[" + str(b) + "," + str(p) + "]")


# NEW tracking van de locaties van schepen
sp_var = {}

for v in range(len(V)):
    for r in range(len(R)): #NEW now also keeps track of location while sailing
        for p in range(len(P)):
            sp_var[v, r, p] = BM.addVar(lb=0, vtype=GRB.BINARY, name="sp_var[" + str(v) + "," + str(r) + "," + str(p) + "]")

# NEW variable pv_var
pv_var = {}
for v in range(len(V)):
    for p in range(len(P)):
        pv_var[v,p] = BM.addVar(lb=0, ub=pv_max[v], vtype=GRB.CONTINUOUS, name="pv_var[" + str(v) + "," + str(p) + "]")

# NEW speed of vessels
speed = {}
for v in range(len(V)):
    for p in range(len(P)):
        speed[v,p] = BM.addVar(lb=0, ub=ts[v], vtype=GRB.CONTINUOUS, name="speed[" + str(v) + "," + str(p) + "]")

# NEW maximum speed of vessels
speed_aux = {}
for v in range(len(V)):
    for r in range(len(R)):
        for p in range(len(P)):
            speed_aux[v,r,p] = BM.addVar(lb=0, vtype=GRB.BINARY, name="speed_aux[" + str(v) + "," + str(r) + "," + str(p) + "]")

# NEW reach of vessels
reach = {}
for v in range(len(V)):
    for p in range(len(P)):
        reach[v,p] = BM.addVar(lb=0, ub=M3, vtype=GRB.CONTINUOUS, name="reach[" + str(v) + "," + str(p) + "]")

# NEW number of vessels at a terminal
nvt = {}
for t in range(len(T)):
    for p in range(len(P)):
        nvt[t,p] = BM.addVar(lb=0, ub=len(V), vtype=GRB.CONTINUOUS, name="nvt[" + str(t) + "," + str(p) + "]")

#NEW - optional, if optional constraint XX is not possible - penalty variable
nvt_aux = {}
for t in range(len(T)):
    for p in range(len(P)):
        nvt_aux[t,p] = BM.addVar(lb=0, ub=(len(V)-1), vtype=GRB.CONTINUOUS, name="nvt_aux[" + str(t) + "," + str(p) + "]")

#NEW - sp_var helper variable to avoid quadratic constraints
sp_var_aux = {}
for v in range(len(V)):
    for r1 in range(len(R)):
        for p in range(len(P)):
            for r2 in range(len(R)):
                sp_var_aux[v, r1, r2, p] = BM.addVar(lb=0, vtype=GRB.BINARY, name="sp_var_aux[" + str(v) + "," + str(r1) + "," + str(r2) + "," + str(p) + "]")

BM.update()
print(u)

# ----------- Objective ---------------------------------------------------------
#NEW minimizing total cost
obj1 = (quicksum(u[t]*cds[t] for t in range(len(T)))
        + quicksum(n[b]*cb[b] for b in range(len(B)))
        + quicksum(CC_p[p] * beta * quicksum(pw_var[b,t,p] for b in range(len(B)))
                   for t in range(len(T)) for p in range(len(P))))

#NEW (optional, if optional constraint XX can't be enforced) - keeps track of busy terminals
obj2 = quicksum(nvt_aux[t,p] for t in range(len(T)) for p in range(len(P)))

BM.setObjective(obj1, GRB.MINIMIZE)
BM.setObjectiveN(obj2, GRB.MINIMIZE, priority=1, name="busy terminals")
BM.update()

# ------------ Constraints ------------------------------------------------------
# Constraint 1 - Used batteries
con1 = {}
for b in range(len(B)):
    BM.addConstr(n[b] * M1, GRB.GREATER_EQUAL, quicksum(quicksum(x[b, t, p] for t in range(len(T)))
        for p in range(1, len(P))) + quicksum(quicksum(y[b,v,p] for v in range(len(V))) for p in range(1, len(P))))

# Constraint 2 - Used terminals
con2 = {}
for t in range(len(T)):
    BM.addConstr(u[t] * M1, GRB.GREATER_EQUAL, quicksum(quicksum(x[b,t,p]*n[b] for b in range(len(B))) for p in range(len(P))))

# Constraint 3 - Maximum number of used batteries in vessel
con3a = {}
for v in range(len(V)):
    for p in range(1, len(P)):
        BM.addConstr(quicksum(m[b, v, p] for b in range(len(B))), GRB.LESS_EQUAL, 1)

con3b = {}
for v in range(len(V)):
    for p in range(1, len(P)):
        BM.addConstr(quicksum(m[b, v, p] for b in range(len(B))), GRB.GREATER_EQUAL, pv_var[v, p] * M2)

# UPDATED Constraint 4 - Battery level
con4 = {}
for b in range(len(B)):
    for p in range(1, len(P)):
        BM.addConstr(l[b, p], GRB.EQUAL, l[b, p - 1] +(quicksum(pw_var[b,t,p] * k[b, t, p] for t in range(len(T))) - quicksum(m[b, v, p] * pv_var[v, p] for v in range(len(V)))))

# NEW Constraint 4b - NEW Batteries do not drain while at a terminal
con4b = {}
for b in range(len(B)):
    for t in range(len(T)):
        for p in range(1, len(P)):
            BM.addConstr(l[b,p] - l[b,p-1] <= M4 * (1 - x[t,v,p]))
            BM.addConstr(l[b,p] - l[b,p-1] >= -M4 * (1 - x[t,v,p]))

# Constraint 5 - Maximum & minimum battery level
con5a = {}
for b in range(len(B)):
    for p in range(1, len(P)):
        BM.addConstr(l[b, p], GRB.LESS_EQUAL, cap[b] * n[b])

con5b = {}
for b in range(len(B)):
    for p in range(1, len(P)):
        BM.addConstr(l[b, p], GRB.GREATER_EQUAL, cap[b] * mbl * n[b])

# Constraint 6 - Minimum battery level to be placed on a ship
#con6 = {}
#for b in range(len(B)):
#    for v in range(len(V)):
#        for p in range(1, len(P)):
#            BM. addConstr((y[b,v,p]-y[b,v,p -1]) * mlv * cap[b], GRB . LESS_EQUAL, l[b,p])

#TODO-deze weer aanzetten?

# Constraint 7 - Battery location
con7 = {}
for b in range(len(B)):
    for p in range(1, len(P)):
        BM.addConstr(quicksum(y[b, v, p] for v in range(len(V))) + quicksum(x[b, t, p] for t in range(len(T))),
                     GRB.EQUAL, n[b])

# Constraint 8 - Batteries & vessels routes
con8a = {}
for b in range(len(B)):
    for t in range(len(T)):
        for p in range(1, len(P)):
            BM.addConstr(x[b, t, p], GRB.LESS_EQUAL,
                         x[b, t, p - 1] + quicksum(y[b, v, p - 1] * sp_var[v, t, p] for v in range(len(V))))

con8b = {}
for b in range(len(B)):
    for v in range(len(V)):
        for p in range(1, len(P)):
            BM.addConstr(y[b, v, p], GRB.LESS_EQUAL,
                         y[b, v, p - 1] + quicksum(x[b, t, p - 1] * sp_var[v, t, p - 1] for t in range(len(T))))

# Constraint 9 - Maximum number of batteries at DS
con9 = {}
for t in range(len(T)):
    for p in range(1, len(P)):
        BM.addConstr(quicksum(x[b, t, p] for b in range(len(B))), GRB.LESS_EQUAL, cht[t])

# Constraint 10 - Maximum number of batteries on vessel
con10 = {}
for v in range(len(V)):
    for p in range(1, len(P)):
        BM.addConstr(quicksum(y[b, v, p] for b in range(len(B))), GRB.LESS_EQUAL, chv[v])

# Constraint 11 - Battery charging at DS only if it is at DS
con11 = {}
for b in range(len(B)):
    for t in range(len(T)):
        for p in range(1, len(P)):
            BM.addConstr(k[b, t, p], GRB.LESS_EQUAL, x[b, t, p])

# Constraint 12 - Battery used in vessel only if it is on vessel
con12 = {}
for b in range(len(B)):
    for v in range(len(V)):
        for p in range(1, len(P)):
            BM.addConstr(m[b, v, p], GRB.LESS_EQUAL, y[b, v, p])

# Constraint 13 - Initial battery location & battery usage (on vessel and terminal)
con13a = {}
for b in range(len(B)):
    for t in range(len(T)):
        BM.addConstr(x[b, t, 0], GRB.EQUAL, Ut0[b, t])

con13b = {}
for b in range(len(B)):
    for v in range(len(V)):
        BM.addConstr(y[b, v, 0], GRB.EQUAL, Uv0[b, v])

con13c = {}
for b in range(len(B)):
    for t in range(len(T)):
        BM.addConstr(k[b, t, 0], GRB.EQUAL, k0[b, t])

con13d = {}
for b in range(len(B)):
    for v in range(len(V)):
        BM.addConstr(m[b, v, 0], GRB.EQUAL, m0[b, v])

# UPDATE - niet meer nodig
#con14 = {}
#for b in range(len(B)):
#    BM.addConstr(l[b, 0], GRB.LESS_EQUAL, Lb0[b])

# Constraint 15 (optional) - Fix battery or DS
#con15a = {}
#BM.addConstr(n[1], GRB.EQUAL, 1)  # Example fixing battery 2 of the input file to be part of the output fleet

#con15b = {}
#BM.addConstr(u[1], GRB.EQUAL, 1)  # Example fixing terminal 2 of the input file as DS

# Constraint 16a - NEW als batterij opgeladen wordt, kan pw_var niet 0 zijn
for b in range(len(B)):
    for t in range(len(T)):
        for p in range(len(P)):
            BM.addConstr(pw_var[b,t,p], GRB.GREATER_EQUAL, k[b,t,p])

# Constraint 16b - NEW als batterij niet opgeladen wordt, is pw_var 0
for b in range(len(B)):
    for t in range(len(T)):
        for p in range(len(P)):
            BM.addConstr(pw_var[b,t,p] * M5, GRB.LESS_EQUAL, k[b,t,p])

# Constraint 17 - NEW making sure battery charge is the same from beginning to end

for b in range(len(B)):
    BM.addConstr(l[b, 1], GRB.LESS_EQUAL, l[b,len(P)-1])

# Constraint 18 - NEW initial conditions ship location

for v in range(len(V)):
    for r in range(len(R)):
        BM.addConstr(sp_var[v, r, 0], GRB.EQUAL, sp[v,r,0])

# Constraint 19 - NEW constraint arrival time and destination

for v in range(len(V)):
    for r in range(len(R)):
        for p in range(len(P)):
            BM.addConstr(sp_var[v, r, p], GRB.GREATER_EQUAL, LT[v, r, p])

# Constraint 20 - NEW defining speed of vessels
for v in range(len(V)):
    for p in range(len(P)):
        BM.addConstr(speed[v,p] * pv_max[v], GRB.EQUAL, ts[v] * pv_var[v,p])

# Constraint 21 - NEW defining the reach of ships
for v in range(len(V)):
    for r in range(len(R)):
        for p in range(1, len(P)):
            if r < len(T): #reach = 0 if ship is at a terminal
                BM.addConstr(reach[v,p] * sp_var[v,r,p], GRB.EQUAL, 0)
            else:           #reach is gained when 'travelling' at a node
                BM.addConstr(reach[v,p] * sp_var[v,r,p], GRB.EQUAL, (reach[v,p-1] + speed[v,p]) * sp_var[v,r,p])

#for v in range(len(V)): #poging tot linearisatie, werkt nog niet
#    for r in range(len(R)):
#        for p in range(1, len(P)):
#            if r < len(T):
#                BM.addConstr(reach[v, p] <= M3 * (1 - sp_var[v, r, p]))
#            else:
#                BM.addConstr(reach_aux[v, p], GRB.EQUAL, reach[v, p - 1] + speed[v, p])
#                BM.addConstr(reach[v, p], GRB.EQUAL, reach_aux[v, p] * sp_var[v, r, p])

# Constraint 22 - NEW initial reach
for v in range(len(V)):
    BM.addConstr(reach[v,0],GRB.EQUAL,(1 - quicksum(sp_var[v,t,0] for t in range(len(T)))) * speed[v,0])

# Constraint XX - NEW initial sp_var_aux
for v in range(len(V)):
    for r1 in range(len(R)):
        for r2 in range(len(R)):
            BM.addConstr(sp_var_aux[v,r1,r2,0], GRB.EQUAL, 0)
#TODO- maak sp_var_aux 1 p korter en pas de constraints aan, scheelt variabelen, dan kan dit ook weg

#Constraint 23 - NEW travelling from terminals to connected nodes and preventing direct travel between terminals

#for v in range(len(V)):
#    for t in range(len(T)):
#        for p in range(1, len(P)):
#            for r in range(len(R)):
#                BM.addConstr(DM[t,r] * sp_var[v,t,p-1] * sp_var[v,r,p], GRB.LESS_EQUAL, M3 - 1)

for v in range(len(V)): #Linear reduces number quadratic constraints from 17949 to 13349
    for t in range(len(T)):
        for p in range(1, len(P)):
            for r in range(len(R)):
                BM.addConstr(sp_var_aux[v,t,r,p], GRB.LESS_EQUAL, sp_var[v,t,p-1])
                BM.addConstr(sp_var_aux[v,t,r,p], GRB.LESS_EQUAL, sp_var[v,r,p])
                BM.addConstr(sp_var_aux[v,t,r,p], GRB.GREATER_EQUAL, sp_var[v,t,p-1] + sp_var[v,r,p] - 1)
                BM.addConstr(DM[t, r] * sp_var_aux[v, t, r, p], GRB.LESS_EQUAL, M3 - 1)

#Constraint 24a - NEW travelling from nodes to terminals
#for v in range(len(V)):
#    for n1 in range(len(T), len(R)):
#        for p in range(1, len(P)):
#            for t in range(len(T)):
#                BM.addConstr(DM[n1,t] * sp_var[v,n1,p-1] * sp_var[v,t,p], GRB.LESS_EQUAL, M3 - 1)
#                BM.addConstr(reach[v,p-1], GRB.GREATER_EQUAL, DM[n1,t] * sp_var[v,n1,p-1] * sp_var[v,t,p])

for v in range(len(V)): #Linear version further removes number q constraints from 13349 to 8749
    for n1 in range(len(T),len(R)):
        for p in range(1, len(P)):
            for t in range(len(T)):
                BM.addConstr(sp_var_aux[v, n1, t, p], GRB.LESS_EQUAL, sp_var[v, n1, p - 1])
                BM.addConstr(sp_var_aux[v, n1, t, p], GRB.LESS_EQUAL, sp_var[v, t, p])
                BM.addConstr(sp_var_aux[v, n1, t, p], GRB.GREATER_EQUAL, sp_var[v, n1, p - 1] + sp_var[v, t, p] - 1)
                BM.addConstr(DM[n1, t] * sp_var_aux[v, n1, t, p], GRB.LESS_EQUAL, M3 - 1)
                BM.addConstr(reach[v, p - 1], GRB.GREATER_EQUAL, DM[n1, t] * sp_var_aux[v, n1, t, p])


#TODO- onderzoek potentiele synergy met sp_var_aux[v,r1,r2,p] in linear constraint 23 en 24!!

#Constraint 24b - NEW a vessel is always at a node or terminal
for v in range(len(V)):
    for p in range(len(P)):
        BM.addConstr(quicksum(sp_var[v,r,p] for r in range(len(R))), GRB.EQUAL, 1)

# Constraint 25 - NEW ships can't travel between nodes
#for v in range(len(V)):
#    for n1 in range(len(T), len(R)):
#        for p in range(1, len(P)):
#            for n2 in range(len(T), len(R)):
#                BM.addConstr(DM[n1, n2] * sp_var[v,n1,p] * sp_var[v,n2,p+1], GRB.LESS_EQUAL,0)

for v in range(len(V)): #Linear version further reduces number qconstraints from 8749 to 6449
    for n1 in range(len(T), len(R)):
        for p in range(1, len(P)):
            for n2 in range(len(T), len(R)):
                BM.addConstr(sp_var_aux[v, n1, n2, p], GRB.LESS_EQUAL, sp_var[v, n1, p - 1])
                BM.addConstr(sp_var_aux[v, n1, n2, p], GRB.LESS_EQUAL, sp_var[v, n2, p])
                BM.addConstr(sp_var_aux[v, n1, n2, p], GRB.GREATER_EQUAL, sp_var[v, n1, p - 1] + sp_var[v, n2, p] - 1)

                BM.addConstr(DM[n1, n2] * sp_var_aux[v,n1,n2,p], GRB.LESS_EQUAL,0)

#Constraint 26 - NEW pv_var must come from batteries, if pv_var>0, 1 battery is used (together with constraint 3)
for v in range(len(V)):
    for p in range(len(P)):
        BM.addConstr(quicksum(m[b, v, p] for b in range(len(B))), GRB.GREATER_EQUAL, M2 * pv_var[v, p])

# Constraint 27a - defining speed_aux
#for v in range(len(V)):
#    for r in range(len(R)):
#        for p in range(1, len(P)):
#            BM.addConstr(speed_aux[v,r,p], GRB.EQUAL, sp_var[v,r,p] * sp_var[v,r,p-1])

for v in range(len(V)): # Linear version reduces n qconstraints from 6449 to 5529
    for r in range(len(R)):
        for p in range(1, len(P)):
            BM.addConstr(speed_aux[v,r,p], GRB.LESS_EQUAL, sp_var[v, r, p])
            BM.addConstr(speed_aux[v,r,p], GRB.LESS_EQUAL, sp_var[v, r, p - 1])
            BM.addConstr(speed_aux[v,r,p], GRB.GREATER_EQUAL, sp_var[v, r, p] + sp_var[v, r, p-1] - 1)


# Constraint 27b - defining initial helper variable speed_aux
for v in range(len(V)):
    for r in range(len(R)):
        BM.addConstr(speed_aux[v,r,0], GRB.EQUAL, 0)

# Constraint 28 - enforcing steady sailing
for v in range(len(V)):
    for n1 in range(len(T), len(R)):
        for p in range(1, len(P)):
            BM.addConstr(speed_aux[v,n1,p] * speed[v, p], GRB.EQUAL,
                         speed_aux[v,n1,p] * speed[v, p - 1])

# Constraint 29 - number of schips at terminal t at time p
for t in range(len(T)):
    for p in range(len(P)):
        BM.addConstr(nvt[t,p], GRB.EQUAL, quicksum(sp_var[v,t,p] for v in range(len(V))))

# Optional Constraint 30 (use if allowed) - only one ship can be at a terminal at one time
#for t in range(len(T)):
#    for p in range(len(P)):
#        BM.addConstr(nvt[t,p], GRB.LESS_EQUAL, 1)

# Optional Constriant 31 (if above is not possible) - count when more than one ship at terminal
for t in range(len(T)):
    for p in range(len(P)):
        BM.addConstr(nvt_aux[t,p], GRB.GREATER_EQUAL, (nvt[t,p] - 1))

# -----------Solve -------------------------------------------------------------
BM.setParam('OutputFlag', True)  # show the Gurobi output
BM.setParam('MIPGap', 0)  # find the optimal solution
BM.setParam('OBBT', 2) # NEW: make solver more aggressive (bound tightening)
#BM.setParam('WorkLimit', 400) #NEW: maximum amount of work (iterations)
#BM.setParam('TimeLimit', 20) #NEW: maximum amount of time
BM.setParam('MIPFocus', 2) #NEW changing high-level solution strategy
#BM.setParam('Threads', 1) #TEST - limit threads from 20 to X
#BM.setParam('NodefileStart', 1) #NEW saving RAM by saving to disk more often
BM.write("output.lp")  # print the model in .lp format file

BM.optimize()


# -----------Print -------------------------------------------------------------
## Batteries
batteries = 0
for i0 in n.values():
    if i0.X > 0:
         batteries = batteries + 1

v_names = np.array([i[0].value for i in wb.worksheets[2]['C2':'C5']])
batt_vessel = []
for b in range(len(B)):
    for v in range(len(V)):
        usage = sum(m[b, v, p].X for p in range(1, len(P)))
        if usage > 0:
            batt_vessel.append([b, v_names[v]])

## Docking stations
ds_names = np.array([i[0].value for i in wb.worksheets[1]['E2':'E6']])
DS = 0
uL = []
Aux1 = []
for i2 in u.values():
    if i2.X > 0:
        DS = DS + 1
        Aux1.append(1)
    else:
        Aux1.append(0)

for i3 in range(len(Aux1)):
    if Aux1[i3] > 0:
        uL.append(ds_names[i3])

## KPIs
Aux2 = 0
for i3 in n.values():
    if i3.X > 0:
        Aux2 = Aux2 + 1
bv = Aux2 / len(V)

Aux3 = 0
for i4 in u.values():
    if i4.X > 0:
        Aux3 = Aux3 + 1
dsv = Aux3 / len(V)

Aux4 = []
for i5 in n.values():
    if i5.X > 0:
        Aux4.append(1)
    else:
        Aux4.append(0)

Aux5 = []
for i6 in u.values():
    if i6.X > 0:
        Aux5.append(1)
    else:
        Aux5.append(0)


cxv = (sum(Aux4 * cb) + sum(Aux5 * cds)) / len(V)
vxv = (obj1.getValue() - cxv * len(V)) / len(V) #NEW variable expense per vessel

pw_list = [] # NEW - list of charging per hour
for p in range(len(P)):
    pw_current = sum(pw_var[b,t,p] for t in range(len(T)) for b in range(len(B)))
    pw_list.append(pw_current.getValue())
pw_list = np.array(pw_list)
pw_list = pw_list / max(pw_list)# normalize


cheapness_list = max(CC_p) - CC_p
cheapness_list = cheapness_list / max(cheapness_list)

corr = np.corrcoef(pw_list, cheapness_list)[0, 1]
R2 = corr ** 2

plt.plot(P, cheapness_list, label = 'affordable energy')
plt.plot(P, pw_list, label = 'total charging')
plt.legend()
plt.show()

print('\n--------------------------------------------------------------------\n')
print('OBTAINED RESULTS \n')
print('')
print('Used terminals as docking stations:', DS)
print('')
print('Docking stations locations:', uL)
print('')
print('')
print('Used batteries:', batteries)
print('')
print('Shared batteries:\n')
print(tabulate(batt_vessel, headers=["Battery #", "Vessel"]))
print('')
print('')
print('KPIs:\n')
print(' * Batteries / vessel:', bv)
print('')
print(' * DS/ vessel:', dsv)
print('')
print(' * Capex / vessel:', cxv)
print('')
print('* Varex / vessel over', (len(P)/24) * beta / 365.25 , 'jaar:', vxv)
print('')
print('* Varex / vessel per jaar:', vxv / jaren)
print('')
print('R2 between charging en variable energy costs:', R2)
print('')
print('Pearson Correlation between charging en variable energy costs:', corr)

print('\n--------------------------------------------------------------------\n')
# ----------- Save outputs in Excel file ----------------------------------------
nInfo = []
for a1 in n.values():
    if a1.X > 0:
        nInfo.append(a1.varName)
uInfo = []
for a2 in u.values():
    if a2.X > 0:
        uInfo.append(a2.varName)
xInfo = []
for a3 in x.values():
    if a3.X > 0:
        xInfo.append(a3.varName)
yInfo = []
for a4 in y.values():
    if a4.X > 0:
        yInfo.append(a4.varName)
lName = []
lValue = []
for a5 in l.values():
    if a5.X > 0:
        lName.append(a5.varName)
        lValue.append(a5.X)
kInfo = []
for a6 in k.values():
    if a6.X > 0:
        kInfo.append(a6.varName)
mInfo = []
for a7 in m.values():
    if a7.X > 0:
        mInfo.append(a7.varName)
spInfo = [] #NEW variable sailing profile
for a8 in sp_var.values():
    if a8.X > 0:
        spInfo.append(a8.varName)

pvName = []
pvValue = []
for a9 in pv_var.values():
    if a9.X > 0:
        pvName.append(a9.varName)
        pvValue.append(a9.X)

reachName = []
reachValue = []
for a10 in reach.values():
    reachName.append(a10.varName)
    reachValue.append(a10.X)

# Create dataframes for each variable
df1 = pd.DataFrame({'n[b]': nInfo})
df2 = pd.DataFrame({'u[t]': uInfo})
df3 = pd.DataFrame({'x[b,t,p]': xInfo})
df4 = pd.DataFrame({'y[b,v,p]': yInfo})
df5 = pd.DataFrame({'l[b,p]': lName, 'Level': lValue})
df6 = pd.DataFrame({'k[b,t,p]': kInfo})
df7 = pd.DataFrame({'m[b,v,p]': mInfo})
df8 = pd.DataFrame({'sp_var[v,t,p]': spInfo})
df9 = pd.DataFrame({'pv[v,p]': pvName, 'pv': pvValue})
df10 = pd.DataFrame({'reach[v,p]': reachName, 'reach': reachValue})

with pd.ExcelWriter('Output.xlsx', engine='xlsxwriter') as writer:
    df1.to_excel(writer, sheet_name='Batteries n(b)')
    df2.to_excel(writer, sheet_name='Docking stations u(t)')
    df3.to_excel(writer, sheet_name='Batt. loc. terminal x(b,t,p)')
    df4.to_excel(writer, sheet_name='Batt. loc. vessel y(b,v,p)')
    df5.to_excel(writer, sheet_name='Batt. level (b,p)')
    df6.to_excel(writer, sheet_name='Batt. ch. terminal k(b,t,p)')
    df7.to_excel(writer, sheet_name='Batt. used vessel m(b,v,p)')
    df8.to_excel(writer, sheet_name='sp_var')
    df9.to_excel(writer, sheet_name='pv_var')
    df10.to_excel(writer, sheet_name='reach')

# ----------- Plotting info ----------------------------------------

# Create figure and 3D axes
fig = plt.figure()
ax = fig.add_subplot(111, projection='3d')
ax.view_init(azim=103) #was (azim=103)

# Iterate through the tensor and plot the binary values
for v in range(len(V)):
    for r in range(len(R)):
        for p in range(len(P)):
            if sp_var[v, r, p].X == 1:
                if r < len(T):
                    ax.scatter(v + 1, r + 1, p + 1, color='b', marker='o')  # Plot blue points for binary value 1
                else:
                    ax.scatter(v + 1, r + 1, p + 1, color='g', marker='o')  # Plot blue points for binary value 1
            else:
                ax.scatter(v + 1, r + 1, p + 1, color='r', marker='x')  # Plot red points for binary value 0
#for v in range(len(V)):
#    for r in range(len(R)):
#        for p in range(len(P)):
#            if LT[v, r, p] == 1:
#                ax.scatter(v + 1, r + 1, p + 1, color='b', marker='o')  # Plot blue points for binary value 1
#            else:
#                ax.scatter(v + 1, r + 1, p + 1, color='r', marker='x')  # Plot red points for binary value 0

# Set labels and title
ax.set_xlabel('v')
ax.set_ylabel('r')
ax.set_zlabel('p')
ax.set_title('3D Binary Tensor Visualization')

# Show plot
plt.show()

def plot_ship_timeline(BM, V, P, B, cap, l, speed, reach, sp_var):
    num_ships = len(V)
    num_timesteps = len(P)

    for v in range(num_ships):
        fig, axs = plt.subplots(4, 1, figsize=(10, 8), sharex=True)
        fig.suptitle(f'Timeline for Ship {V[v]}')

        # Battery levels
        battery_levels = np.zeros((len(B), num_timesteps))
        for b in range(len(B)):
            for p in range(num_timesteps):
                battery_levels[b, p] = y[b, v, p].x * l[b, p].x  # Extract variable value

        # Use a colormap to get a list of colors
        colormap = plt.cm.get_cmap('tab20', len(B))
        colors = [colormap(i) for i in range(len(B))]

        for b in range(len(B)):
            axs[0].scatter(range(num_timesteps), battery_levels[b], label=f'Battery {B[b]} (cap: {cap[b]} kW)', color=colors[b], marker='o')
        axs[0].set_ylabel('Charge of batteries on boat')
        axs[0].grid()
#        axs[0].legend(loc='upper right')

        # Speed
        ship_speeds = np.zeros(num_timesteps)
        for p in range(num_timesteps):
            ship_speeds[p] = speed[v, p].x  # Extract variable value

        axs[1].plot(range(num_timesteps), ship_speeds, label='Speed', marker='.')
        axs[1].set_ylabel('Speed')
        axs[1].grid()
#        axs[1].legend(loc='upper right')

        # Reach
        ship_reach = np.zeros(num_timesteps)
        for p in range(num_timesteps):
            ship_reach[p] = reach[v, p].x  # Extract variable value

        axs[2].plot(range(num_timesteps), ship_reach, label='Reach', marker='.')
        axs[2].set_ylabel('Reach')
        axs[2].grid()
#        axs[2].legend(loc='upper right')

        # Location
        ship_locations = np.zeros(num_timesteps)
        for p in range(num_timesteps):
            for r in range(len(R)):
                if sp_var[v, r, p].x > 0.5:  # Extract variable value
                    ship_locations[p] = r

        axs[3].scatter(range(num_timesteps), ship_locations, label='Location', marker='o')
        axs[3].set_ylabel('Location')
        axs[3].set_xlabel('Time Step')
        axs[3].grid()
#        axs[3].legend(loc='upper right')

        plt.tight_layout(rect=[0, 0.03, 1, 0.95])
        plt.show()

# Example call after the model optimization
plot_ship_timeline(BM, V, P, B, cap, l, speed, reach, sp_var)



busy_terminals = np.zeros((len(T), len(P)))
for t in range(len(T)):
    for p in range(len(P)):
        busy_terminals[t, p] = nvt[t,p].x

for t in range(len(T)):
    fig, axs = plt.subplots(1, 1, figsize=(10, 8), sharex=True)
    fig.suptitle(f'Timeline for Terminal {T[t]}')

    axs.plot(range(len(P)), busy_terminals[t,:], label='Busy Terminal', marker='o')
    axs.set_ylabel('Number of vessels at terminal')
    axs.grid()


    plt.tight_layout(rect=[0, 0.03, 1, 0.95])
    plt.show()
